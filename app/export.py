"""Export a generated report to a single, manager-ready Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
FLAG_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
TITLE_FONT = Font(size=14, bold=True)


def _fmt_date(period):
    start, end = period
    return f"{start.strftime('%b %d, %Y')} - {end.strftime('%b %d, %Y')}"


def _write_header_row(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _autofit(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _pct(value):
    return None if value is None else round(value * 100, 1)


def _write_summary_sheet(wb, report):
    ws = wb.active
    ws.title = 'Summary'

    ws['A1'] = 'Weekly Sales Summary'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = _fmt_date(report['current_period'])
    if report['has_prior_data']:
        ws['A2'] = ws['A2'].value + f"  (vs. {_fmt_date(report['prior_period'])})"

    ws['A4'] = report['paragraph']
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A4:F8')

    row = 10
    _write_header_row(ws, row, ['Metric', 'This Period', 'Change vs. Prior Period'])
    metrics = [
        ('Total Sales', report['current_totals']['sales'], report['totals_change']['sales']),
        ('Total Profit', report['current_totals']['profit'], report['totals_change']['profit']),
        ('Total Orders', report['current_totals']['orders'], report['totals_change']['orders']),
    ]
    for i, (label, value, change) in enumerate(metrics, start=1):
        r = row + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=round(value, 2))
        ws.cell(row=r, column=3, value=f"{_pct(change)}%" if change is not None else 'N/A')

    flag_row = row + len(metrics) + 2
    ws.cell(row=flag_row, column=1, value=f"Flagged Issues ({len(report['flags'])})").font = Font(bold=True)
    for i, flag in enumerate(report['flags'], start=1):
        r = flag_row + i
        cell = ws.cell(row=r, column=1, value=_flag_text(flag))
        cell.fill = FLAG_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    _autofit(ws, [22, 16, 22, 14, 14, 14])


def _flag_text(flag):
    kind = flag['kind']
    if kind == 'margin':
        return (f"{flag['name']}: margin down {_pct(abs(flag['change']))}% vs. typical "
                f"({_pct(flag['margin_now'])}% now vs {_pct(flag['margin_typical'])}% typical)")
    label = {'region': 'Region', 'category': 'Category', 'store': 'Store', 'product': 'Product'}[kind]
    return f"{label} — {flag['name']}: sales down {_pct(abs(flag['change']))}% vs. prior period"


def _write_breakdown_sheet(wb, title, rows, extra_columns=None):
    ws = wb.create_sheet(title)
    headers = ['Name', 'Sales', 'Profit', 'Orders', 'Prior Sales', 'Change vs. Prior']
    _write_header_row(ws, 1, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r['name'])
        ws.cell(row=i, column=2, value=round(r['sales'], 2))
        ws.cell(row=i, column=3, value=round(r['profit'], 2))
        ws.cell(row=i, column=4, value=r['orders'])
        ws.cell(row=i, column=5, value=round(r['prior_sales'], 2) if r['prior_sales'] is not None else None)
        change_cell = ws.cell(row=i, column=6, value=f"{_pct(r['sales_change'])}%" if r['sales_change'] is not None else 'N/A')
        if r['sales_change'] is not None and r['sales_change'] <= -0.10:
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = FLAG_FILL
    _autofit(ws, [24, 16, 16, 12, 16, 18])
    return ws


def _write_discount_sheet(wb, discount_rows):
    ws = wb.create_sheet('Discounted Products')
    headers = ['Product', 'Avg. Discount', 'Sales', 'Margin Now', 'Typical Margin',
               'Margin Change', 'Profit Impact vs. Prior Week']
    _write_header_row(ws, 1, headers)
    for i, r in enumerate(discount_rows, start=2):
        ws.cell(row=i, column=1, value=r['product'])
        ws.cell(row=i, column=2, value=f"{_pct(r['avg_discount'])}%")
        ws.cell(row=i, column=3, value=round(r['sales'], 2))
        ws.cell(row=i, column=4, value=f"{_pct(r['margin_now'])}%")
        ws.cell(row=i, column=5, value=f"{_pct(r['margin_typical'])}%" if r['margin_typical'] is not None else 'N/A')
        ws.cell(row=i, column=6, value=f"{_pct(r['margin_change'])}%" if r['margin_change'] is not None else 'N/A')
        ws.cell(row=i, column=7, value=r['impact_label'])
        if r['margin_change'] is not None and r['margin_change'] <= -0.15:
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = FLAG_FILL
    _autofit(ws, [24, 14, 16, 14, 14, 14, 24])


def export_report_to_excel(report, path):
    """Write the full report to a single .xlsx file at `path`."""
    wb = Workbook()
    _write_summary_sheet(wb, report)
    _write_breakdown_sheet(wb, 'By Region', report['region_rows'])
    _write_breakdown_sheet(wb, 'By Category', report['category_rows'])
    _write_breakdown_sheet(wb, 'By Store', report['store_rows'])
    _write_discount_sheet(wb, report['discount_rows'])
    wb.save(path)
